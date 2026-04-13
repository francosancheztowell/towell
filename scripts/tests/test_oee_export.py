import unittest
from unittest.mock import patch
from datetime import date
from typing import cast

from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from scripts import oee_export


class MergeSyncTests(unittest.TestCase):
    def test_sync_after_row_insert_rebuilds_without_incremental_merge_operations(self):
        wb = Workbook()
        ws = cast(Worksheet, wb.active)
        ws.merge_cells("A1:B1")
        ws.merge_cells("C5:D6")
        ws.merge_cells("E8:F9")

        ws.insert_rows(5, 2)

        with (
            patch.object(
                ws.merged_cells,
                "remove",
                side_effect=AssertionError("expected bulk merge rebuild"),
            ),
            patch.object(
                ws.merged_cells,
                "add",
                side_effect=AssertionError("expected bulk merge rebuild"),
            ),
        ):
            oee_export.sync_merged_ranges_after_row_insert(ws, 5, 2)

        self.assertEqual(
            {str(rng) for rng in ws.merged_cells.ranges},
            {"A1:B1", "C7:D8", "E10:F11"},
        )

    def test_sync_after_row_delete_rebuilds_without_incremental_merge_operations(self):
        wb = Workbook()
        ws = cast(Worksheet, wb.active)
        ws.merge_cells("A1:B1")
        ws.merge_cells("C7:D8")
        ws.merge_cells("E10:F11")

        ws.delete_rows(5, 2)

        with (
            patch.object(
                ws.merged_cells,
                "remove",
                side_effect=AssertionError("expected bulk merge rebuild"),
            ),
            patch.object(
                ws.merged_cells,
                "add",
                side_effect=AssertionError("expected bulk merge rebuild"),
            ),
        ):
            oee_export.sync_merged_ranges_after_row_delete(ws, 5, 2)

        self.assertEqual(
            {str(rng) for rng in ws.merged_cells.ranges},
            {"A1:B1", "C5:D6", "E8:F9"},
        )

    def test_clear_section_values_can_skip_second_unmerge_pass(self):
        wb = Workbook()
        ws = cast(Worksheet, wb.active)
        ws["A1"] = "keep"
        ws["B2"] = "clear"

        with patch.object(
            oee_export,
            "unmerge_rows_in_range",
            side_effect=AssertionError("unexpected unmerge pass"),
        ):
            oee_export.clear_section_values(
                ws,
                {"top": 1, "footer": 2},
                already_unmerged=True,
            )

        self.assertIsNone(ws["A1"].value)
        self.assertIsNone(ws["B2"].value)

    def test_write_section_data_can_skip_merge_cleanup_on_fresh_section(self):
        wb = Workbook()
        ws = cast(Worksheet, wb.active)
        layout = oee_export.build_layout([], section_top_row=1)
        layout["week_number"] = 13

        with patch.object(
            oee_export,
            "_unmerge_col_range",
            side_effect=AssertionError("unexpected merge cleanup"),
        ):
            actual_rows = oee_export.write_section_data(
                ws,
                {"top": 1, "footer": layout["footer_row"]},
                layout,
                date(2026, 3, 23),
                {},
                clear_existing_merges=False,
            )

        self.assertEqual(actual_rows, [])
        self.assertIn("A2:A8", {str(rng) for rng in ws.merged_cells.ranges})


class CopyToDestinationTests(unittest.TestCase):
    @staticmethod
    def _os_error(*, winerror=None, errno_value=None):
        err = OSError("copy failed")
        if winerror is not None:
            err.winerror = winerror
        if errno_value is not None:
            err.errno = errno_value
        return err

    @staticmethod
    def _permission_error(*, winerror=None, errno_value=None):
        err = PermissionError("permission denied")
        if winerror is not None:
            err.winerror = winerror
        if errno_value is not None:
            err.errno = errno_value
        return err

    def test_copy_result_retries_on_unexpected_network_error(self):
        with (
            patch.object(
                oee_export.shutil,
                "copy2",
                side_effect=[self._os_error(winerror=59), None],
            ) as copy2_mock,
            patch.object(oee_export.time, "sleep") as sleep_mock,
            patch.object(oee_export, "write_status") as status_mock,
        ):
            result = oee_export.copy_oee_result_to_destination(
                "src.xlsx",
                "dst.xlsx",
                "status.json",
                0.0,
                "tok",
            )

        self.assertIsNone(result)
        self.assertEqual(copy2_mock.call_count, 2)
        sleep_mock.assert_called_once()
        status_mock.assert_called_once()

    def test_copy_result_returns_sibling_file_when_main_replace_stays_blocked(self):
        retry_errors = [self._os_error(winerror=32) for _ in range(15)]

        with (
            patch.object(
                oee_export.shutil,
                "copy2",
                side_effect=[*retry_errors, None],
            ) as copy2_mock,
            patch.object(
                oee_export.os, "replace", side_effect=self._os_error(winerror=32)
            ),
            patch.object(oee_export.time, "sleep") as sleep_mock,
            patch.object(oee_export, "write_status") as status_mock,
        ):
            result = oee_export.copy_oee_result_to_destination(
                "src.xlsx",
                r"C:\share\OEE_ATADORES.xlsx",
                "status.json",
                0.0,
                "tok",
            )

        self.assertIsNotNone(result)
        self.assertIn("OEE_ATADORES_NUEVO_tok.xlsx", result)
        self.assertEqual(copy2_mock.call_count, 16)
        self.assertEqual(sleep_mock.call_count, 14)
        self.assertEqual(status_mock.call_count, 15)

    def test_copy_result_raises_when_sibling_replace_fails_for_non_retryable_reason(
        self,
    ):
        retry_errors = [self._os_error(winerror=32) for _ in range(15)]

        with (
            patch.object(
                oee_export.shutil,
                "copy2",
                side_effect=[*retry_errors, None],
            ) as copy2_mock,
            patch.object(
                oee_export.os,
                "replace",
                side_effect=self._os_error(winerror=5, errno_value=13),
            ),
            patch.object(oee_export.time, "sleep") as sleep_mock,
            patch.object(oee_export, "write_status") as status_mock,
        ):
            with self.assertRaises(OSError):
                oee_export.copy_oee_result_to_destination(
                    "src.xlsx",
                    r"C:\share\OEE_ATADORES.xlsx",
                    "status.json",
                    0.0,
                    "tok",
                )

        self.assertEqual(copy2_mock.call_count, 16)
        self.assertEqual(sleep_mock.call_count, 14)
        self.assertEqual(status_mock.call_count, 15)

    def test_copy_result_does_not_retry_non_retryable_permission_error(self):
        with (
            patch.object(
                oee_export.shutil,
                "copy2",
                side_effect=self._permission_error(winerror=5, errno_value=13),
            ) as copy2_mock,
            patch.object(oee_export.time, "sleep") as sleep_mock,
            patch.object(oee_export, "write_status") as status_mock,
        ):
            with self.assertRaises(PermissionError):
                oee_export.copy_oee_result_to_destination(
                    "src.xlsx",
                    "dst.xlsx",
                    "status.json",
                    0.0,
                    "tok",
                )

        self.assertEqual(copy2_mock.call_count, 1)
        sleep_mock.assert_not_called()
        status_mock.assert_not_called()

    def test_copy_source_to_local_temp_retries_on_transient_network_error(self):
        with (
            patch.object(
                oee_export.shutil,
                "copy2",
                side_effect=[self._os_error(winerror=59), None],
            ) as copy2_mock,
            patch.object(oee_export.time, "sleep") as sleep_mock,
            patch.object(oee_export, "write_status") as status_mock,
        ):
            oee_export.copy_oee_source_to_local_temp(
                r"\\server\share\OEE_ATADORES.xlsx",
                r"C:\Temp\oee.xlsx",
                "status.json",
                0.0,
            )

        self.assertEqual(copy2_mock.call_count, 2)
        sleep_mock.assert_called_once()
        status_mock.assert_called_once()


class StyleRegressionTests(unittest.TestCase):
    def test_apply_row_style_restores_snapshot_fill(self):
        wb = Workbook()
        ws = cast(Worksheet, wb.active)
        ws["A1"].fill = PatternFill("solid", fgColor="FF0000")

        snapshot = oee_export.snapshot_section_styles(ws, {"top": 1, "footer": 1})

        ws["A1"].fill = PatternFill("solid", fgColor="00FF00")
        oee_export._apply_row_style(ws, 1, snapshot[0])

        self.assertEqual(ws["A1"].fill.patternType, "solid")
        self.assertEqual(ws["A1"].fill.fgColor.rgb, "00FF0000")

    def test_apply_ck_cu_styles_keeps_calidad_numeric_and_time_columns_as_duration(
        self,
    ):
        wb = Workbook()
        ws = cast(Worksheet, wb.active)

        oee_export._apply_ck_cu_styles(ws, 1, [(4, "AT1")])

        self.assertEqual(ws.cell(4, 91).number_format, "General")
        for col in (92, 93, 94, 95):
            self.assertEqual(ws.cell(4, col).number_format, "[h]:mm:ss")

    def test_apply_ck_cu_styles_keeps_aggregate_rows_on_fixed_summary_offsets(self):
        wb = Workbook()
        ws = cast(Worksheet, wb.active)

        oee_export._apply_ck_cu_styles(ws, 1, [(6, "AT1")])

        agg_fill = oee_export._ck_styles()["agg_fill"].fgColor.rgb
        self.assertEqual(ws.cell(11, 89).fill.fgColor.rgb, agg_fill)
        self.assertEqual(ws.cell(13, 89).fill.fgColor.rgb, agg_fill)
        self.assertEqual(ws.cell(14, 89).fill.fgColor.rgb, agg_fill)

    def test_write_ck_cu_formulas_keeps_headers_for_empty_week(self):
        wb = Workbook()
        ws = cast(Worksheet, wb.active)
        layout = oee_export.build_layout([], section_top_row=1)

        rows = oee_export._write_ck_cu_formulas(ws, layout, 1, {})

        self.assertEqual(rows, [])
        self.assertEqual(ws.cell(3, 89).value, "CVE")
        self.assertEqual(ws.cell(3, 90).value, "ATADOR")

    def test_write_ck_cu_formulas_skips_reserved_aggregate_rows(self):
        wb = Workbook()
        ws = cast(Worksheet, wb.active)
        layout = {
            "turns": {
                1: {
                    "blocks": [
                        {"atador_key": f"AT{i}", "row_start": 100 + i}
                        for i in range(1, 9)
                    ]
                }
            }
        }

        rows = oee_export._write_ck_cu_formulas(ws, layout, 1, {})
        summary_rows = [row for row, _ in rows]

        self.assertNotIn(11, summary_rows)
        self.assertNotIn(13, summary_rows)
        self.assertNotIn(14, summary_rows)

    def test_write_ck_cu_formulas_clamps_reserved_rows_to_last_detail_row(self):
        wb = Workbook()
        ws = cast(Worksheet, wb.active)
        layout = oee_export.build_layout([], section_top_row=1)
        layout["turns"][1]["blocks"][0]["atador_key"] = "AT1"

        oee_export._write_ck_cu_formulas(ws, layout, 1, {})

        agg_fill = oee_export._ck_styles()["agg_fill"].fgColor.rgb
        self.assertEqual(ws.cell(11, 89).fill.fgColor.rgb, agg_fill)
        self.assertNotEqual(ws.cell(13, 89).fill.fgColor.rgb, agg_fill)


class SemanaLabelPatchTests(unittest.TestCase):
    def test_patch_semana_sheet_does_not_touch_defined_plain_text_header(self):
        wb = Workbook()
        ws = cast(Worksheet, wb.active)
        ws.cell(3, 5).value = "TIEMPO DE ATADO"
        ws.cell(3, 6).value = "Semana 04: del 20/01/2026 al 26/01/2026"

        oee_export.patch_semana_sheet_tiempo_week_label(ws, 5, 2026)

        self.assertEqual(ws.cell(3, 6).value, "Semana 04: del 20/01/2026 al 26/01/2026")

    def test_patch_semana_sheet_does_not_touch_defined_formula_header(self):
        wb = Workbook()
        ws = cast(Worksheet, wb.active)
        ws.cell(3, 5).value = "TIEMPO DE ATADO"
        ws.cell(3, 6).value = "=DETALLE!H2"

        oee_export.patch_semana_sheet_tiempo_week_label(ws, 5, 2026)

        self.assertEqual(ws.cell(3, 6).value, "=DETALLE!H2")


if __name__ == "__main__":
    unittest.main()
