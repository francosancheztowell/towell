#!/usr/bin/env python3
"""
OEE Atadores Excel Export — Python replacement for the PHP PhpSpreadsheet job.

Worker (misma máquina que `php artisan queue:work` en cola oee-atadores):
  pip install -r scripts/requirements.txt
  ODBC Driver 17 or 18 for SQL Server (pyodbc).

Config Laravel: config/oee.php — OEE_PYTHON_BINARY, OEE_PYTHON_TIMEOUT, OEE_EXPORT_DB_CONNECTION.

Usage:
    python oee_export.py \
        --week-start 2026-03-23 \
        --week-end   2026-03-23 \
        --token      abc123 \
        --file-path  "\\\\192.168.2.11\\ti-system\\OEE_ATADORES.xlsx" \
        --status-file "C:\\xampp\\htdocs\\Towell\\storage\\app\\oee_python_abc123.json" \
        --db-host HOST --db-database DB --db-username USER --db-password PASS
"""

from __future__ import annotations

import argparse
import errno
import json
import os
import random
import re
import shutil
import sys
import tempfile
import time
import traceback
from copy import copy
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from typing import Any

# ---------------------------------------------------------------------------
# Optional dotenv support (graceful if not installed)
# ---------------------------------------------------------------------------
try:
    from dotenv import load_dotenv as _load_dotenv

    def _try_load_dotenv() -> None:
        # Walk up from script location looking for .env
        current = Path(__file__).resolve().parent
        for _ in range(5):
            env_file = current / ".env"
            if env_file.is_file():
                _load_dotenv(env_file)
                return
            current = current.parent
except ImportError:

    def _try_load_dotenv() -> None:
        pass


# ---------------------------------------------------------------------------
# Constants (mirror of Reporte00EAtadoresExport.php)
# ---------------------------------------------------------------------------
DETAIL_START_ROW = 4  # 1-indexed, relative to section top
BASE_DETAIL_ROWS = 42
FOOTER_BASE_ROW = 46
DEFAULT_BLOCK_HEIGHT = 6
MIN_BLOCKS_PER_TURN = 2
CAPACITACION_HEIGHT = 6
MAX_COLUMN_INDEX = 88  # CJ

# Python-only: detalle más compacto que el export PHP (DEFAULT_BLOCK_HEIGHT / CAPACITACION_HEIGHT arriba).
COMPACT_DEFAULT_BLOCK_HEIGHT = 1
COMPACT_CAPACITACION_HEIGHT = 3

# Turn definitions (relative prototype rows, 1-indexed within section)
TURN_DEFINITIONS = {
    1: {"label": "1 ER TURNO", "first_prototype": [4, 9], "next_prototype": [10, 15]},
    2: {"label": "2 DO TURNO", "first_prototype": [16, 21], "next_prototype": [22, 27]},
    3: {"label": "3 ER TURNO", "first_prototype": [28, 33], "next_prototype": [34, 39]},
}

# Column definitions per day (0 = Monday)
DAY_DEFINITIONS = {
    0: {
        "header": "H2",
        "start": "D",
        "end": "E",
        "duration": "F",
        "avg_time": "G",
        "atado": "H",
        "telar": "I",
        "calif": "J",
        "five_s": "K",
        "avg_calif": "L",
        "merma": "M",
        "avg_merma": "N",
        "footer_count": "I",
    },
    1: {
        "header": "T2",
        "start": "P",
        "end": "Q",
        "duration": "R",
        "avg_time": "S",
        "atado": "T",
        "telar": "U",
        "calif": "V",
        "five_s": "W",
        "avg_calif": "X",
        "merma": "Y",
        "avg_merma": "Z",
        "footer_count": "U",
    },
    2: {
        "header": "AF2",
        "start": "AB",
        "end": "AC",
        "duration": "AD",
        "avg_time": "AE",
        "atado": "AF",
        "telar": "AG",
        "calif": "AH",
        "five_s": "AI",
        "avg_calif": "AJ",
        "merma": "AK",
        "avg_merma": "AL",
        "footer_count": "AG",
    },
    3: {
        "header": "AR2",
        "start": "AN",
        "end": "AO",
        "duration": "AP",
        "avg_time": "AQ",
        "atado": "AR",
        "telar": "AS",
        "calif": "AT",
        "five_s": "AU",
        "avg_calif": "AV",
        "merma": "AW",
        "avg_merma": "AX",
        "footer_count": "AS",
    },
    4: {
        "header": "BD2",
        "start": "AZ",
        "end": "BA",
        "duration": "BB",
        "avg_time": "BC",
        "atado": "BD",
        "telar": "BE",
        "calif": "BF",
        "five_s": "BG",
        "avg_calif": "BH",
        "merma": "BI",
        "avg_merma": "BJ",
        "footer_count": "BE",
    },
    5: {
        "header": "BP2",
        "start": "BL",
        "end": "BM",
        "duration": "BN",
        "avg_time": "BO",
        "atado": "BP",
        "telar": "BQ",
        "calif": "BR",
        "five_s": "BS",
        "avg_calif": "BT",
        "merma": "BU",
        "avg_merma": "BV",
        "footer_count": "BQ",
    },
    6: {
        "header": "CB2",
        "start": "BX",
        "end": "BY",
        "duration": "BZ",
        "avg_time": "CA",
        "atado": "CB",
        "telar": "CC",
        "calif": "CD",
        "five_s": "CE",
        "avg_calif": "CF",
        "merma": "CG",
        "avg_merma": "CH",
        "footer_count": "CC",
    },
}

# Summary columns (CK..CU) mirrors writeCkCuFormulas
AVG_CALIF_COLS = ["L", "X", "AJ", "AV", "BH", "BT", "CF"]
AVG_TIME_COLS = ["G", "S", "AE", "AQ", "BC", "BO", "CA"]
AVG_MERMA_COLS = ["N", "Z", "AL", "AX", "BJ", "BV", "CH"]

# ---------------------------------------------------------------------------
# Status file helpers
# ---------------------------------------------------------------------------

# Última fase del export (para mensajes de error cuando la excepción no trae texto).
_RUN_PHASE = "inicio"

# ---------------------------------------------------------------------------
# CK-CU summary table styles (cols 89-99, lazy-initialized)
# ---------------------------------------------------------------------------
_CK_STYLES: dict | None = None

_CK_HDR_LABELS = [
    "CVE",
    "ATADOR",
    "CALIDAD",
    "TIEMPO",
    "ESTANDAR",
    "DIFER.",
    "T.AJUST.",
    "EFIC.%",
    "CALIF.%",
    "MERMA",
    "ECO",
]

_RETRYABLE_COPY_WINERRORS = {32, 33, 59, 64, 121, 1231, 1232}
_RETRYABLE_COPY_ERRNOS = {errno.EACCES, errno.EPERM, errno.EBUSY}


def _is_retryable_copy_error(exc: BaseException) -> bool:
    winerror = getattr(exc, "winerror", None)
    errno_ = getattr(exc, "errno", None)

    if winerror is not None:
        return winerror in _RETRYABLE_COPY_WINERRORS

    return errno_ in _RETRYABLE_COPY_ERRNOS


def _copy_retry_delay(attempt: int) -> float:
    return min(50.0, 2.5 * (1.55 ** (attempt - 1))) + random.uniform(0.0, 2.5)


def copy_oee_source_to_local_temp(
    src: str,
    dst: str,
    status_file: str,
    elapsed_start: float,
) -> None:
    """Copy the source workbook from SMB/local storage to a temp file with retries."""
    max_retries = 15
    last_exc: BaseException | None = None

    for attempt in range(1, max_retries + 1):
        try:
            shutil.copy2(src, dst)
            return
        except (PermissionError, OSError) as exc:
            last_exc = exc
            if not _is_retryable_copy_error(exc):
                raise
            if attempt >= max_retries:
                break
            delay = _copy_retry_delay(attempt)
            write_status(
                status_file,
                "procesando",
                "Esperando acceso al archivo origen en red (puede estar abierto en Excel o haber un fallo transitorio SMB). "
                f"Intento {attempt}/{max_retries}, siguiente reintento en ~{round(delay)} s...",
                time.time() - elapsed_start,
            )
            time.sleep(delay)

    raise PermissionError(
        "No se pudo copiar el archivo origen tras "
        f"{max_retries} intentos. Cierra OEE_ATADORES.xlsx en Excel si esta abierto y revisa la conexion al share. "
        f"Ultimo error: {last_exc!r}"
    ) from last_exc


def _ck_styles() -> dict:
    global _CK_STYLES  # noqa: PLW0603
    if _CK_STYLES is None:
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side  # noqa: PLC0415

        thin = Side(style="thin")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        _CK_STYLES = {
            "hdr_fill": PatternFill("solid", fgColor="2F5496"),
            "data_fill": PatternFill("solid", fgColor="DCE6F1"),
            "agg_fill": PatternFill("solid", fgColor="BDD7EE"),
            "hdr_font": Font(bold=True, color="FFFFFF", name="Calibri", size=9),
            "body_font": Font(name="Calibri", size=9),
            "border": border,
            "center": Alignment(horizontal="center"),
        }
    return _CK_STYLES


def _set_run_phase(phase: str) -> None:
    global _RUN_PHASE  # noqa: PLW0603
    _RUN_PHASE = phase


def _format_exception_detail(exc: BaseException) -> str:
    """Texto útil cuando str(exc) es vacío o solo tuplas (p. ej. algunos errores ODBC)."""
    parts: list[str] = [f"{type(exc).__name__}"]
    s = str(exc).strip()
    if s:
        parts.append(s)
    if getattr(exc, "args", None):
        dec: list[Any] = []
        for a in exc.args:
            if isinstance(a, (bytes, bytearray)):
                dec.append(bytes(a).decode("utf-8", "replace"))
            else:
                dec.append(a)
        parts.append(f"args={dec!r}")
    if isinstance(exc, OSError):
        for name in ("winerror", "errno", "strerror", "filename", "filename2"):
            v = getattr(exc, name, None)
            if v is not None:
                parts.append(f"{name}={v!r}")
    return " | ".join(parts)[:2000]


def write_status(
    status_file: str, estado: str, mensaje: str = "", elapsed_s: float = 0
) -> None:
    """Write job status JSON atomically. Never raises."""
    try:
        payload = {
            "estado": estado,
            "mensaje": mensaje,
            "elapsed_s": round(elapsed_s, 1),
            "updated_at": datetime.now(timezone.utc).isoformat(),
        }
        tmp = status_file + ".tmp"
        with open(tmp, "w", encoding="utf-8") as f:
            json.dump(payload, f)
        os.replace(tmp, status_file)
    except Exception as exc:  # noqa: BLE001
        print(
            f"CRITICAL: cannot write status file {status_file}: {exc}", file=sys.stderr
        )


def copy_oee_result_to_destination(
    src: str, dst: str, status_file: str, elapsed_start: float, token: str
) -> str | None:
    """
    Copy finished workbook to the network path. WinError 32 is common when Excel
    (or preview/indexer) holds OEE_ATADORES.xlsx open — retry with backoff.

    Returns None if the main destination file was updated successfully.
    Returns a str path if the export was saved to a sibling file (same folder) because
    the main file stayed locked — caller should mention this in the completion message.

    Raises PermissionError / OSError if neither direct copy nor sibling fallback works.
    """

    max_retries = 15
    last_exc: BaseException | None = None
    for attempt in range(1, max_retries + 1):
        try:
            shutil.copy2(src, dst)
            return None
        except (PermissionError, OSError) as exc:
            last_exc = exc
            if not _is_retryable_copy_error(exc):
                raise
            if attempt >= max_retries:
                break
            delay = _copy_retry_delay(attempt)
            write_status(
                status_file,
                "procesando",
                "Esperando acceso al archivo en red (puede estar abierto en Excel o haber un fallo transitorio SMB). "
                f"Intento {attempt}/{max_retries}, siguiente reintento en ~{round(delay)} s…",
                time.time() - elapsed_start,
            )
            time.sleep(delay)

    # Fallback: escribir con nombre nuevo en la misma carpeta
    # y sustituir al principal con os.replace; si el principal sigue bloqueado, queda la copia.
    dst_path = Path(dst)
    safe_tok = re.sub(r"[^\w.-]+", "_", token)[:24] or "export"
    sibling = dst_path.parent / f"{dst_path.stem}_NUEVO_{safe_tok}{dst_path.suffix}"

    write_status(
        status_file,
        "procesando",
        "Archivo bloqueado: guardando copia alternativa en la misma carpeta…",
        time.time() - elapsed_start,
    )

    try:
        shutil.copy2(src, sibling)
    except OSError as exc:
        raise PermissionError(
            "No se pudo copiar al archivo principal tras "
            f"{max_retries} intentos ni crear una copia alternativa en la misma carpeta. "
            "Cierra OEE_ATADORES.xlsx en Excel (todas las PCs con el archivo abierto) y reintenta. "
            f"Error al principal: {last_exc!r}. Error alternativa: {exc!r}"
        ) from exc

    try:
        os.replace(str(sibling), dst)
        return None
    except OSError as exc:
        if not _is_retryable_copy_error(exc):
            raise
        # La copia nueva existe y contiene el export correcto; el usuario puede renombrar cuando libere el principal.
        return str(sibling)


# ---------------------------------------------------------------------------
# Argument parsing
# ---------------------------------------------------------------------------


def parse_args() -> argparse.Namespace:
    _try_load_dotenv()

    p = argparse.ArgumentParser(description="OEE Atadores Excel Export")
    p.add_argument("--week-start", required=True)
    p.add_argument("--week-end", required=True)
    p.add_argument("--token", required=True)
    p.add_argument("--file-path", required=True)
    p.add_argument("--status-file", required=True)
    p.add_argument("--db-host", default=os.environ.get("DB_HOST", "localhost"))
    p.add_argument("--db-database", default=os.environ.get("DB_DATABASE", ""))
    p.add_argument("--db-username", default=os.environ.get("DB_USERNAME", ""))
    p.add_argument("--db-password", default=os.environ.get("DB_PASSWORD", ""))
    p.add_argument(
        "--db-port", default=int(os.environ.get("DB_PORT", "1433")), type=int
    )
    p.add_argument("--log-file", default=None)
    return p.parse_args()


# ---------------------------------------------------------------------------
# Database
# ---------------------------------------------------------------------------


def _odbc_server_clause(host: str, port: int) -> str:
    """
    Instancia con nombre (host\\INSTANCIA): no añadir ,puerto — ODBC falla si se mezcla.
    Solo host o IP: usar host,puerto como siempre.
    """
    h = (host or "").strip()
    if not h:
        return f"localhost,{port}"
    if "\\" in h:
        return h
    return f"{h},{port}"


def get_db_connection(args: argparse.Namespace):
    import pyodbc  # noqa: PLC0415

    drivers = [d for d in pyodbc.drivers() if "SQL Server" in d]
    if not drivers:
        raise RuntimeError(
            "No ODBC SQL Server driver found. Install 'ODBC Driver 17 for SQL Server' from Microsoft."
        )
    driver = next((d for d in drivers if "17" in d or "18" in d), drivers[0])

    server = _odbc_server_clause(args.db_host, args.db_port)
    conn_str = (
        f"DRIVER={{{driver}}};"
        f"SERVER={server};"
        f"DATABASE={args.db_database};"
        f"UID={args.db_username};"
        f"PWD={args.db_password};"
        "Encrypt=no;TrustServerCertificate=yes;"
        "MARS_Connection=yes;"
    )
    try:
        return pyodbc.connect(conn_str, timeout=30)
    except Exception as e:  # noqa: BLE001
        raise RuntimeError(
            "No se pudo conectar a SQL Server por ODBC. "
            "Revisa DB_HOST (si usas instancia con \\\\ no pongas puerto en .env), usuario, contraseña y firewall. "
            f"Detalle: {e!r}"
        ) from e


def query_records(conn, week_start_str: str, week_end_str: str) -> list[dict]:
    """Mirror of PHP preloadRecordsByWeek — fetches all records in range."""
    # Range end = last day of the last requested week (sunday)
    week_start = date.fromisoformat(week_start_str)
    week_end = date.fromisoformat(week_end_str)
    range_end = week_end + timedelta(days=6)

    sql = """
        SELECT Id, FechaArranque, Turno, CveTejedor, NomTejedor, Tipo,
               NoTelarId, HrInicio, HoraArranque, Calidad, Limpieza, MergaKg
        FROM dbo.AtaMontadoTelas
        WHERE Estatus = 'Autorizado'
          AND FechaArranque IS NOT NULL
          AND FechaArranque >= ? AND FechaArranque <= ?
        ORDER BY FechaArranque, Turno, CveTejedor, NomTejedor, HrInicio, HoraArranque, Id
    """
    cursor = conn.cursor()
    try:
        cursor.execute(sql, week_start_str, str(range_end))
    except Exception as e:  # noqa: BLE001
        raise RuntimeError(
            "Fallo al ejecutar consulta en dbo.AtaMontadoTelas "
            f"(FechaArranque entre {week_start_str!r} y {str(range_end)!r}): "
            f"{_format_exception_detail(e)}"
        ) from e
    cols = [col[0] for col in cursor.description]
    rows = []
    for row in cursor.fetchall():
        rows.append(dict(zip(cols, row)))
    cursor.close()
    return rows


# ---------------------------------------------------------------------------
# Record preparation (mirrors PHP prepareRecords)
# ---------------------------------------------------------------------------


def _monday_of(d: date) -> date:
    return d - timedelta(days=d.weekday())


def normalize_turn(value) -> int | None:
    if value is None:
        return None
    try:
        t = int(str(value).strip())
        return t if t in (1, 2, 3) else None
    except (ValueError, TypeError):
        return None


def resolve_atador_key(row: dict) -> str:
    cve = str(row.get("CveTejedor") or "").strip()
    if cve:
        return cve
    return str(row.get("NomTejedor") or "").strip()


def resolve_atado_label(tipo) -> str:
    t = str(tipo or "").strip().lower()
    if t == "rizo":
        return "R"
    if t == "pie":
        return "P"
    return t[0].upper() if t else ""


def normalize_time(value) -> str | None:
    if value is None:
        return None
    raw = str(value).strip()
    if not raw:
        return None
    m = re.match(r"^(\d{1,2}:\d{2}:\d{2})", raw)
    if m:
        return m.group(1)
    m = re.match(r"^(\d{1,2}:\d{2})$", raw)
    if m:
        return m.group(1) + ":00"
    # try datetime parse
    for fmt in ("%H:%M:%S", "%H:%M", "%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S"):
        try:
            return datetime.strptime(raw, fmt).strftime("%H:%M:%S")
        except ValueError:
            continue
    return None


def normalize_number(value):
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (ValueError, TypeError):
        return None


def prepare_records(raw_rows: list[dict], week_monday: date) -> list[dict]:
    """Mirror of PHP prepareRecords."""
    out = []
    for row in raw_rows:
        fa = row.get("FechaArranque")
        if fa is None:
            continue
        if isinstance(fa, datetime):
            fecha = fa.date()
        elif isinstance(fa, date):
            fecha = fa
        else:
            try:
                fecha = datetime.fromisoformat(str(fa)).date()
            except ValueError:
                continue

        turno = normalize_turn(row.get("Turno"))
        if turno is None:
            continue

        day_index = (fecha - week_monday).days
        if day_index < 0 or day_index > 6:
            continue

        out.append(
            {
                "id": int(row.get("Id") or 0),
                "fecha": fecha,
                "day_index": day_index,
                "turno": turno,
                "atador_key": resolve_atador_key(row),
                "atado": resolve_atado_label(row.get("Tipo")),
                "telar": str(row.get("NoTelarId") or "").strip(),
                "hora_inicio": normalize_time(row.get("HrInicio")),
                "hora_fin": normalize_time(row.get("HoraArranque")),
                "calif": normalize_number(row.get("Calidad")),
                "five_s": normalize_number(row.get("Limpieza")),
                "merma": normalize_number(row.get("MergaKg")),
            }
        )
    return out


def group_records_by_week(raw_rows: list[dict]) -> dict[str, list[dict]]:
    """Group raw rows by ISO Monday date string."""
    groups: dict[str, list[dict]] = {}
    for row in raw_rows:
        fa = row.get("FechaArranque")
        if fa is None:
            continue
        if isinstance(fa, datetime):
            fecha = fa.date()
        elif isinstance(fa, date):
            fecha = fa
        else:
            try:
                fecha = datetime.fromisoformat(str(fa)).date()
            except ValueError:
                continue
        monday = str(_monday_of(fecha))
        groups.setdefault(monday, []).append(row)
    return groups


# ---------------------------------------------------------------------------
# Layout builder (mirrors PHP buildLayout)
# ---------------------------------------------------------------------------


def build_layout(records: list[dict], section_top_row: int = 1) -> dict:
    """
    Build section layout mirroring PHP buildLayout().
    section_top_row: 1-indexed absolute row where this section starts in the sheet.
    Returns dict with turns, capacitacion, footer_row, week_number, detail_start_row, last_detail_row.
    """

    def offset(row: int) -> int:
        return section_top_row + row - 1

    # Group by turn → atador_key
    by_turn: dict[int, dict[str, list[dict]]] = {}
    for rec in records:
        by_turn.setdefault(rec["turno"], {}).setdefault(rec["atador_key"], []).append(
            rec
        )

    turns = {}
    current_row = offset(DETAIL_START_ROW)

    for turno, tdef in TURN_DEFINITIONS.items():
        atador_map = by_turn.get(turno, {})
        sorted_keys = sorted(atador_map.keys(), key=str.casefold)

        blocks = []
        for atador_key in sorted_keys:
            items = atador_map[atador_key]
            items_by_day: list[list[dict]] = [[] for _ in range(7)]
            for item in items:
                items_by_day[item["day_index"]].append(item)
            max_daily = max((len(d) for d in items_by_day), default=0)
            blocks.append(
                {
                    "atador_key": atador_key,
                    "items_by_day": items_by_day,
                    "height": max(COMPACT_DEFAULT_BLOCK_HEIGHT, max_daily),
                }
            )

        while len(blocks) < MIN_BLOCKS_PER_TURN:
            blocks.append(
                {
                    "atador_key": "",
                    "items_by_day": [[] for _ in range(7)],
                    "height": COMPACT_DEFAULT_BLOCK_HEIGHT,
                }
            )

        turn_start = current_row
        for i, block in enumerate(blocks):
            proto = tdef["first_prototype"] if i == 0 else tdef["next_prototype"]
            block["row_start"] = current_row
            block["row_end"] = current_row + block["height"] - 1
            block["prototype"] = [offset(proto[0]), offset(proto[1])]
            current_row = block["row_end"] + 1

        turns[turno] = {
            "label": tdef["label"],
            "row_start": turn_start,
            "row_end": current_row - 1,
            "blocks": blocks,
        }

    cap_start = current_row
    cap_end = cap_start + COMPACT_CAPACITACION_HEIGHT - 1
    footer_row = cap_end + 1

    return {
        "turns": turns,
        "capacitacion": {
            "label": "CAPACITACION",
            "row_start": cap_start,
            "row_end": cap_end,
            "prototype": [offset(40), offset(45)],
        },
        "detail_start_row": offset(DETAIL_START_ROW),
        "last_detail_row": cap_end,
        "footer_row": footer_row,
    }


# ---------------------------------------------------------------------------
# Section parsing (mirrors OeeAtadoresFileService::parseDetalleSections)
# ---------------------------------------------------------------------------


def normalize_label(value) -> str:
    if value is None:
        return ""
    return " ".join(str(value).upper().split())


def _build_merged_top_left_cache(ws) -> dict[tuple[int, int], tuple[int, int]]:
    """
    Map (row,col) → (top_row, top_col) para celdas que no son la esquina de un merge.
    Un barrido O(área combinada); el parseo de DETALLE evita O(filas × merges) por celda.
    """
    cache: dict[tuple[int, int], tuple[int, int]] = {}
    for m in ws.merged_cells.ranges:
        tl = (m.min_row, m.min_col)
        for r in range(m.min_row, m.max_row + 1):
            for c in range(m.min_col, m.max_col + 1):
                if (r, c) != tl:
                    cache[(r, c)] = tl
    return cache


def _merged_cell_top_left(
    ws,
    row: int,
    col: int,
    merge_cache: dict[tuple[int, int], tuple[int, int]] | None = None,
) -> tuple[int, int]:
    """
    If (row,col) is inside a merged range, return top-left of that range.
    Fast path: non-top-left merged cells are stored as MergedCell objects in
    ws._cells; regular cells (including top-left of a merge) are not MergedCell.
    Avoids iterating all merged_cells.ranges for the common non-merged case.
    """
    if merge_cache is not None:
        hit = merge_cache.get((row, col))
        if hit is not None:
            return hit

    from openpyxl.cell.cell import MergedCell  # noqa: PLC0415

    cell = ws._cells.get((row, col))
    if cell is None or not isinstance(cell, MergedCell):
        return row, col
    # Cell is a non-top-left merged cell; find its top-left via ranges
    for m in ws.merged_cells.ranges:
        if m.min_row <= row <= m.max_row and m.min_col <= col <= m.max_col:
            return m.min_row, m.min_col
    return row, col


def get_cell_effective_value(
    ws,
    row: int,
    col: int,
    merge_cache: dict[tuple[int, int], tuple[int, int]] | None = None,
):
    """Value for display/parse; resolves merged cells (openpyxl only stores value on top-left)."""
    r, c = _merged_cell_top_left(ws, row, col, merge_cache)
    return ws.cell(r, c).value


def parse_detalle_sections(ws) -> dict:
    """
    Scan DETALLE worksheet for section markers.
    Returns {"sections": [...], "map": {week_num: section}}.
    Each section: {top, start, footer, rows, week} — top = max(1, start_row - 1) per PHP.
    """
    from openpyxl.utils import column_index_from_string as col_idx  # noqa: PLC0415

    COL_A = 1
    COL_B = 2
    COL_CJ = col_idx("CJ")

    merge_cache = _build_merged_top_left_cache(ws)

    max_row = ws.max_row or 1
    starts: list[int] = []
    footers: list[int] = []

    for row in range(1, max_row + 1):
        val_a = normalize_label(get_cell_effective_value(ws, row, COL_A, merge_cache))
        val_b = normalize_label(get_cell_effective_value(ws, row, COL_B, merge_cache))
        val_cj = normalize_label(get_cell_effective_value(ws, row, COL_CJ, merge_cache))

        # One marker per merged region (openpyxl only stores value on top-left; we resolve merges)
        if val_a == "SEMANA":
            tl_r, tl_c = _merged_cell_top_left(ws, row, COL_A, merge_cache)
            if tl_r == row and tl_c == COL_A:
                starts.append(row)

        if val_b == "SEMANA" and val_cj == "ATADOS":
            tl_r, tl_c = _merged_cell_top_left(ws, row, COL_B, merge_cache)
            if tl_r == row and tl_c == COL_B:
                footers.append(row)

    sections: list[dict[str, Any]] = []
    section_map: dict[int, dict] = {}
    footer_index = 0

    for start_row in starts:
        while footer_index < len(footers) and footers[footer_index] <= start_row:
            footer_index += 1

        if footer_index >= len(footers):
            break

        footer_row = footers[footer_index]
        footer_index += 1

        top_row = max(1, start_row - 1)
        week_num = _resolve_week_number(ws, footer_row, merge_cache)

        section = {
            "top": top_row,
            "start": start_row,
            "footer": footer_row,
            "rows": footer_row - top_row + 1,
            "week": week_num,
        }
        sections.append(section)

        if week_num is not None and week_num not in section_map:
            section_map[week_num] = section

    sections.sort(key=lambda s: s["top"])

    return {"sections": sections, "map": section_map}


def _resolve_week_formula_value(
    ws,
    value: Any,
    depth: int = 0,
    merge_cache: dict[tuple[int, int], tuple[int, int]] | None = None,
) -> int | None:
    """
    Mirror PHP resolveWeekFormulaValue for =COLrow or =COLrow+/-N.
    Handles optional delta (=C46, =C46+1, =C46-1) and ignores spaces.
    """
    if depth > 8 or not isinstance(value, str):
        return None
    formula = value.strip().replace(" ", "")
    if not formula.startswith("="):
        return None
    m = re.match(r"^=([A-Z]+)(\d+)([+-]\d+)?$", formula.upper())
    if not m:
        return None
    col_l, row_s, delta_s = m.group(1), m.group(2), m.group(3) or "+0"
    from openpyxl.utils import column_index_from_string as col_idx  # noqa: PLC0415

    r = int(row_s)
    c = col_idx(col_l)
    ref_val = get_cell_effective_value(ws, r, c, merge_cache)
    if isinstance(ref_val, (int, float)) and not isinstance(ref_val, bool):
        base_week = int(ref_val)
    else:
        base_week = _resolve_week_formula_value(ws, ref_val, depth + 1, merge_cache)
    if base_week is None:
        return None
    week = base_week + int(delta_s)
    return week if 1 <= week <= 53 else None


def _resolve_week_number(
    ws,
    footer_row: int,
    merge_cache: dict[tuple[int, int], tuple[int, int]] | None = None,
) -> int | None:
    COL_C = 3
    value = get_cell_effective_value(ws, footer_row, COL_C, merge_cache)
    if value is None:
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        v = int(value)
        return v if 1 <= v <= 53 else None
    raw = str(value).strip()
    if raw.lstrip("-").isdigit():
        v = int(raw)
        return v if 1 <= v <= 53 else None
    return _resolve_week_formula_value(ws, value, 0, merge_cache)


def resolve_insert_top(sections: list[dict], week_num: int) -> int:
    """Mirror OeeAtadoresFileService::resolveInsertTop."""
    for sec in sections:
        w = sec.get("week")
        if w is not None and w > week_num:
            return int(sec["top"])
    if not sections:
        return 1
    last = sections[-1]
    return int(last["footer"]) + 1


def assert_single_iso_year(weeks: list[date]) -> int:
    years = {d.isocalendar()[0] for d in weeks}
    if len(years) != 1:
        raise RuntimeError(
            "El rango debe pertenecer al mismo año ISO para actualizar el archivo anual OEE."
        )
    return years.pop()


def resolve_workbook_year(wb) -> int | None:
    import re  # noqa: PLC0415

    for name in wb.sheetnames:
        if re.match(r"^ATADORES\s+\d{4}$", name.strip()):
            m = re.search(r"(\d{4})$", name.strip())
            if m:
                return int(m.group(1))
    return None


def normalize_detalle_footer_weeks(ws, sections: list[dict]) -> None:
    """Mirror normalizeDetalleFooterWeeks — force numeric week in column C at each footer."""
    for sec in sections:
        w = sec.get("week")
        if w is None:
            continue
        ws.cell(sec["footer"], 3).value = int(w)


def resolve_detalle_visual_week_anchor(ws, section: dict) -> int | None:
    """Mirror resolveDetalleVisualWeekAnchor."""
    top = int(section.get("top", 0))
    footer = int(section.get("footer", 0))
    if top < 1 or footer < top:
        return None

    candidates: list[int] = []
    for m in list(ws.merged_cells.ranges):
        if m.min_col != 1 or m.max_col != 1:
            continue
        row_start, row_end = m.min_row, m.max_row
        if row_start < top or row_end > footer:
            continue
        val = normalize_label(get_cell_effective_value(ws, row_start, 1))
        if val == "" or val == "SEMANA":
            continue
        candidates.append(row_start)

    if candidates:
        return max(candidates)

    for row in range(footer, top - 1, -1):
        val = normalize_label(get_cell_effective_value(ws, row, 1))
        if val == "" or val == "SEMANA":
            continue
        return row

    return None


def normalize_detalle_visual_weeks(ws, sections: list[dict]) -> None:
    """Mirror normalizeDetalleVisualWeeks."""
    for sec in sections:
        w = sec.get("week")
        if w is None:
            continue
        anchor = resolve_detalle_visual_week_anchor(ws, sec)
        if anchor is not None:
            ws.cell(anchor, 1).value = int(w)


# ---------------------------------------------------------------------------
# SEMANA sheets — label next to "TIEMPO DE ATADO" (not updated by DETALLE-only export)
# ---------------------------------------------------------------------------


def _format_semana_week_range_label(week_num: int, iso_year: int) -> str:
    """Human-readable ISO week span (Mon–Sun); avoids Excel showing 01/01/1900 when value is 0."""
    try:
        mon = date.fromisocalendar(iso_year, week_num, 1)
    except ValueError:
        return f"Semana {week_num}"
    sun = mon + timedelta(days=6)
    return f"Semana {week_num}: del {mon.strftime('%d/%m/%Y')} al {sun.strftime('%d/%m/%Y')}"


def _parse_iso_week_from_semana_sheet_title(name: str) -> int | None:
    m = re.match(r"^SEMANA\s*0*(\d{1,2})\s*$", name.strip(), re.IGNORECASE)
    if not m:
        return None
    w = int(m.group(1))
    return w if 1 <= w <= 53 else None


def _parse_iso_week_from_semana_b2(ws) -> int | None:
    raw = ws["B2"].value
    if raw is None:
        return None
    m = re.search(r"(?:SEMANA|Semana)\s*0*(\d{1,2})\b", str(raw).strip(), re.IGNORECASE)
    if not m:
        return None
    w = int(m.group(1))
    return w if 1 <= w <= 53 else None


def _should_patch_tiempo_adjacent_value(value: Any) -> bool:
    """True only for clearly broken header values that should be replaced."""
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return 0.0 <= float(value) < 10.0
    if isinstance(value, date) and value.year <= 1900:
        return True
    return False


def patch_semana_sheet_tiempo_week_label(ws, week_num: int, iso_year: int) -> None:
    """
    Row 3 often has E3='TIEMPO DE ATADO' and the next cell was a =DETALLE!… link to a day header.
    Replace only clearly broken values (0 / 1900 dates) with an explicit week-range
    string. If the sheet already has a defined text or formula header, preserve it.
    """
    label = _format_semana_week_range_label(week_num, iso_year)
    header_row = 3
    for col in range(1, 18):
        if (
            normalize_label(get_cell_effective_value(ws, header_row, col))
            != "TIEMPO DE ATADO"
        ):
            continue
        for next_col in range(col + 1, min(col + 4, 22)):
            cell = ws.cell(header_row, next_col)
            if _should_patch_tiempo_adjacent_value(cell.value):
                cell.value = label
                cell.number_format = "@"
                return
        return


def patch_semana_sheets_week_labels(wb, iso_year: int) -> None:
    """Apply patch_semana_sheet_tiempo_week_label to each 'SEMANA …' worksheet."""
    for name in wb.sheetnames:
        if not re.match(r"^SEMANA\s", name.strip(), re.IGNORECASE):
            continue
        ws = wb[name]
        wn = _parse_iso_week_from_semana_sheet_title(name)
        if wn is None:
            wn = _parse_iso_week_from_semana_b2(ws)
        if wn is None:
            continue
        patch_semana_sheet_tiempo_week_label(ws, wn, iso_year)


# ---------------------------------------------------------------------------
# Cell manipulation helpers
# ---------------------------------------------------------------------------


def to_excel_time(time_str: str) -> float:
    """Convert HH:MM:SS to Excel fractional day."""
    parts = time_str.split(":")
    h, m, s = int(parts[0]), int(parts[1]), int(parts[2]) if len(parts) > 2 else 0
    return (h * 3600 + m * 60 + s) / 86400


def to_excel_date(d: date) -> int:
    """Convert Python date to Excel serial date integer."""
    return (d - date(1899, 12, 30)).days


# ---------------------------------------------------------------------------
# Merges (mirror OeeAtadoresFileService::unmergeRowsInRange)
# ---------------------------------------------------------------------------


def _safe_unmerge(ws, merge_range) -> None:
    """Unmerge a range, tolerating openpyxl's KeyError on orphan MergedCell objects."""
    try:
        ws.unmerge_cells(str(merge_range))
    except KeyError:
        pass


def _cleanup_orphan_merged_cells(ws, start_row: int, end_row: int) -> None:
    """
    After unmerging, openpyxl may leave orphan MergedCell objects in _cells
    because the cleanup loop fails with KeyError for cells never accessed.
    These orphans get shifted by insert_rows into neighboring sections,
    corrupting their structure. Remove them explicitly.
    Iterates only the target row range instead of all cells for performance.
    """
    from openpyxl.cell.cell import MergedCell  # noqa: PLC0415

    orphans = [
        (r, c)
        for r in range(start_row, end_row + 1)
        for c in range(1, MAX_COLUMN_INDEX + 12)
        if isinstance(ws._cells.get((r, c)), MergedCell)
    ]
    for key in orphans:
        del ws._cells[key]


def unmerge_rows_in_range(ws, start_row: int, end_row: int) -> None:
    """
    Remove every merged range that intersects [start_row, end_row],
    then clean up orphan MergedCell objects left in _cells.
    """
    if start_row > end_row:
        return
    to_remove = [
        m
        for m in list(ws.merged_cells.ranges)
        if m.min_row <= end_row and m.max_row >= start_row
    ]
    for m in to_remove:
        _safe_unmerge(ws, m)
    _cleanup_orphan_merged_cells(ws, start_row, end_row)


def _translate_shifted_formulas(ws, before_row: int, amount: int) -> None:
    """
    openpyxl insert_rows / delete_rows moves cell objects in _cells but does NOT
    update the formula strings stored inside those cells.  Template sections for
    weeks that have not yet been processed by this script keep formula-based week
    references (e.g. =C46+1 in col A or footer col C).  After a row shift those
    strings point to stale row numbers; Excel then evaluates them against wrong
    cells and the week label appears blank/wrong.

    This function updates every formula string in the cells that were shifted
    (rows >= before_row after insert, rows that ended up >= before_row after
    delete) so their relative cell references point to the correct new rows.

    Called immediately after insert_rows / delete_rows + the merge sync, before
    any new data is written into those rows.
    """
    if amount == 0:
        return
    try:
        from openpyxl.formula.translate import Translator  # noqa: PLC0415
    except ImportError:
        return

    for (r, c), cell in list(ws._cells.items()):
        if r < before_row:
            continue
        if getattr(cell, "data_type", None) != "f":
            continue
        val = cell.value
        if not isinstance(val, str) or not val.startswith("="):
            continue
        try:
            cell.value = Translator(val, cell.coordinate).translate_formula(
                row_delta=amount, col_delta=0
            )
        except Exception:  # noqa: BLE001
            pass


def _rebuild_shifted_merged_ranges(ws, from_row: int, row_delta: int) -> None:
    """
    Rebuild merged ranges in one pass after rows move.

    Callers must unmerge any range intersecting the edited row band first. This
    helper only shifts ranges that start at or below the row threshold, which
    matches the exporter flow after `unmerge_rows_in_range()` has cleaned the
    active section before insert/delete.
    """
    if row_delta == 0 or not ws.merged_cells.ranges:
        return

    from openpyxl.worksheet.cell_range import CellRange  # noqa: PLC0415

    rebuilt_ranges = set()
    for cr in ws.merged_cells.ranges:
        if cr.min_row >= from_row:
            rebuilt_ranges.add(
                CellRange(
                    min_row=cr.min_row + row_delta,
                    max_row=cr.max_row + row_delta,
                    min_col=cr.min_col,
                    max_col=cr.max_col,
                )
            )
        else:
            rebuilt_ranges.add(cr)

    rebuilt = type(ws.merged_cells)()
    rebuilt.ranges = rebuilt_ranges
    ws.merged_cells = rebuilt


def sync_merged_ranges_after_row_insert(ws, before_row: int, amount: int) -> None:
    """
    openpyxl insert_rows(idx, amount) moves cells in _cells but leaves
    ws.merged_cells stale; shift ranges whose start row is at or below idx.
    Any merge intersecting the inserted band must already be removed by the
    caller (the exporter does this via `unmerge_rows_in_range()`).
    """
    if amount <= 0:
        return
    _rebuild_shifted_merged_ranges(ws, before_row, amount)


def sync_merged_ranges_after_row_delete(ws, start_row: int, amount: int) -> None:
    """
    openpyxl delete_rows(idx, amount) moves cells but not merged_cells; fix ranges
    whose start row is at or below idx + amount. Any merge intersecting the
    deleted band must already be removed by the caller.
    """
    if amount <= 0:
        return
    threshold = start_row + amount
    _rebuild_shifted_merged_ranges(ws, threshold, -amount)


# ---------------------------------------------------------------------------
# Section clearing (values only, preserve styles)
# ---------------------------------------------------------------------------


def clear_section_values(ws, section: dict, already_unmerged: bool = False) -> None:
    """
    Clear values in the section. Unmerge first (like PHP unmergeDynamicRanges)
    so we don't write over inconsistent merged cells.
    """
    top = section["top"]
    footer = section["footer"]
    if not already_unmerged:
        unmerge_rows_in_range(ws, top, footer)
    for row in range(top, footer + 1):
        for col in range(1, 100):
            cell = ws.cell(row, col)
            if cell.value is not None:
                cell.value = None


# ---------------------------------------------------------------------------
# Section resize
# ---------------------------------------------------------------------------


def resize_section(ws, section: dict, desired_rows: int) -> dict:
    """
    Insert or delete rows so section has exactly desired_rows.
    Mirror of PHP resizeDetalleSection: unmerge the section range first,
    then insert/delete. The orphan MergedCell cleanup in unmerge_rows_in_range
    prevents contamination of sections below.
    """
    current_rows = section["rows"]
    delta = desired_rows - current_rows
    if delta == 0:
        return section

    top = section["top"]
    footer = section["footer"]

    unmerge_rows_in_range(ws, top, footer)

    if delta > 0:
        ws.insert_rows(footer, delta)
        sync_merged_ranges_after_row_insert(ws, footer, delta)
        _translate_shifted_formulas(ws, footer, delta)
    else:
        delete_start = footer + delta
        del_count = -delta
        ws.delete_rows(delete_start, del_count)
        sync_merged_ranges_after_row_delete(ws, delete_start, del_count)
        _translate_shifted_formulas(ws, delete_start, -del_count)

    new_footer = footer + delta
    return {
        "top": section["top"],
        "footer": new_footer,
        "rows": desired_rows,
        "week": section["week"],
    }


# ---------------------------------------------------------------------------
# Style snapshot / restore (mirrors PHP copySectionRange + duplicateStyle)
# ---------------------------------------------------------------------------


def snapshot_section_styles(ws, section: dict) -> dict:
    """
    Capture a snapshot of every cell's full style array + row heights within a
    section.
    Returns a dict keyed by relative row offset (0-based from section top).
    Each entry: {"height": float|None, "cells": {col_idx: {"style": ...}}}.
    Using the internal StyleArray is much faster than copying each style
    component (font/fill/border/alignment/protection/format) separately and
    still preserves the full cell formatting when reapplied.
    """
    top = section["top"]
    footer = section["footer"]
    snapshot: dict[int, dict] = {}

    for row in range(top, footer + 1):
        rel = row - top
        rd = ws.row_dimensions.get(row)
        height = rd.height if rd else None

        cells: dict[int, dict] = {}
        for col in range(1, MAX_COLUMN_INDEX + 12):  # up to CU (col 99)
            cell = ws.cell(row, col)
            cells[col] = {"style": copy(cell._style)}

        snapshot[rel] = {"height": height, "cells": cells}

    return snapshot


def _apply_row_style(ws, row: int, style_row: dict) -> None:
    """Apply a captured style_row dict to an absolute worksheet row."""
    h = style_row.get("height")
    if h is not None:
        ws.row_dimensions[row].height = h

    for col, st in style_row.get("cells", {}).items():
        cell = ws.cell(row, col)
        cell._style = copy(st["style"])


def apply_section_styles(ws, section: dict, snapshot: dict, layout: dict) -> None:
    """
    Re-apply captured styles to a section that may have been resized.
    - Rows that existed before resize: re-apply from their original relative offset.
    - New rows (inserted): derive style from prototype rows within the snapshot.
    - Footer row: always takes style from the last row of the original snapshot.
    """
    top = section["top"]
    footer = section["footer"]
    new_total = footer - top + 1
    snap_total = len(snapshot)

    if snap_total == 0:
        return

    # Footer is always last row in original snapshot
    footer_snap_rel = snap_total - 1

    # Build a map: absolute row → relative snapshot row to use as style source
    row_style_map: dict[int, int] = {}

    # Header rows (rows before detail_start_row): map 1:1
    detail_start = layout.get("detail_start_row", top + 3)
    header_count = detail_start - top
    for i in range(min(header_count, snap_total - 1)):
        row_style_map[top + i] = i

    # Footer row: always from original footer
    row_style_map[footer] = footer_snap_rel

    # Capacitacion block (just before footer)
    cap = layout.get("capacitacion", {})
    cap_start = cap.get("row_start", footer - COMPACT_CAPACITACION_HEIGHT)
    cap_end = cap.get("row_end", footer - 1)
    # Original cap started at relative offset snap_total - 1 - COMPACT_CAPACITACION_HEIGHT
    orig_cap_rel_start = footer_snap_rel - COMPACT_CAPACITACION_HEIGHT
    for row in range(cap_start, min(cap_end + 1, footer)):
        cap_offset = row - cap_start
        snap_rel = orig_cap_rel_start + cap_offset
        if 0 <= snap_rel < snap_total:
            row_style_map[row] = snap_rel

    # Turn blocks: use prototype rows from snapshot
    for turn in layout.get("turns", {}).values():
        for block in turn.get("blocks", []):
            proto = block.get("prototype", [])
            if len(proto) < 2:
                continue
            # proto = [abs_proto_start, abs_proto_end] computed against section_top
            proto_start_rel = proto[0] - top
            proto_end_rel = proto[1] - top

            rs = block["row_start"]
            re_ = block["row_end"]
            proto_height = proto_end_rel - proto_start_rel + 1

            for row in range(rs, re_ + 1):
                if row in row_style_map:
                    continue
                block_offset = row - rs
                # Cycle through prototype rows for blocks larger than prototype
                snap_rel = proto_start_rel + (block_offset % proto_height)
                if 0 <= snap_rel < snap_total:
                    row_style_map[row] = snap_rel

    # Apply styles
    for row in range(top, footer + 1):
        snap_rel = row_style_map.get(row)
        if snap_rel is not None and snap_rel in snapshot:
            _apply_row_style(ws, row, snapshot[snap_rel])
        else:
            # Fallback: use nearest known row from snapshot (middle detail row)
            mid = min(snap_total - 2, max(0, header_count))
            if mid in snapshot:
                _apply_row_style(ws, row, snapshot[mid])


# ---------------------------------------------------------------------------
# Section data writer
# ---------------------------------------------------------------------------


def write_section_data(
    ws,
    section: dict,
    layout: dict,
    week_monday: date,
    name_map: dict,
    clear_existing_merges: bool = True,
) -> list[tuple[int, str]]:
    """Write all cell values and formulas into the section.
    Returns list of (detalle_row, atador_key) for use by update_semana_sheet."""
    _write_day_headers(ws, section["top"], week_monday)
    _write_col_a(
        ws,
        layout,
        section["top"],
        layout.get("week_number", week_monday.isocalendar()[1]),
        clear_existing_merges=clear_existing_merges,
    )
    _write_turn_labels(ws, layout, clear_existing_merges=clear_existing_merges)
    _write_atador_blocks(ws, layout, clear_existing_merges=clear_existing_merges)
    _write_footer(ws, layout)
    return _write_ck_cu_formulas(ws, layout, section["top"], name_map)


def _write_day_headers(ws, section_top: int, week_monday: date) -> None:
    """Row section_top+1: Excel serial dates for 7 days."""
    # header cells are like "H2" meaning col H, relative row 2 from section top
    from openpyxl.utils import column_index_from_string as ci  # noqa: PLC0415

    for day_idx, day in DAY_DEFINITIONS.items():
        hdr = day["header"]
        # Parse column letter and relative row from header like "H2", "AF2"
        m = re.match(r"^([A-Z]+)(\d+)$", hdr)
        if not m:
            continue
        col = ci(m.group(1))
        rel_row = int(m.group(2))
        abs_row = section_top + rel_row - 1
        d = week_monday + timedelta(days=day_idx)
        ws.cell(abs_row, col).value = to_excel_date(d)


def _write_col_a(
    ws,
    layout: dict,
    section_top: int,
    week_number: int,
    clear_existing_merges: bool = True,
) -> None:
    """Col A: 'SEMANA' merged over most rows, week_number merged over last Turn3 block."""
    turns = layout["turns"]
    turn3_blocks = turns[3]["blocks"]
    last_t3_block = turn3_blocks[-1]

    header_label_row = section_top + 2 - 1  # relative row 2

    # Unmerge existing A merges in the section range first
    footer_row = layout["footer_row"]
    if clear_existing_merges:
        _unmerge_col_range(ws, 1, 1, section_top, footer_row)

    label_end_row = last_t3_block["row_start"] - 1
    if label_end_row >= header_label_row:
        ws.merge_cells(
            start_row=header_label_row,
            start_column=1,
            end_row=label_end_row,
            end_column=1,
        )
        ws.cell(header_label_row, 1).value = "SEMANA"

    ws.merge_cells(
        start_row=last_t3_block["row_start"],
        start_column=1,
        end_row=last_t3_block["row_end"],
        end_column=1,
    )
    ws.cell(last_t3_block["row_start"], 1).value = week_number


def _unmerge_col_range(
    ws, col_start: int, col_end: int, row_start: int, row_end: int
) -> None:
    """
    Remove any existing merge ranges that overlap the given area.
    Orphan cleanup is intentionally omitted here: this function is called
    many times inside write_section_data (after clear_section_values already
    cleaned all orphans), so the extra full-scan would be wasted work.
    Orphan cleanup is handled by unmerge_rows_in_range before any structural
    row insertions/deletions.
    """
    to_remove = [
        m
        for m in list(ws.merged_cells.ranges)
        if (
            m.min_row <= row_end
            and m.max_row >= row_start
            and m.min_col <= col_end
            and m.max_col >= col_start
        )
    ]
    for m in to_remove:
        _safe_unmerge(ws, m)


def _write_turn_labels(ws, layout: dict, clear_existing_merges: bool = True) -> None:
    """Col B: turn labels merged over full turn rows."""
    for turno, turn in layout["turns"].items():
        # Unmerge first
        if clear_existing_merges:
            _unmerge_col_range(ws, 2, 2, turn["row_start"], turn["row_end"])
        ws.merge_cells(
            start_row=turn["row_start"],
            start_column=2,
            end_row=turn["row_end"],
            end_column=2,
        )
        ws.cell(turn["row_start"], 2).value = turn["label"]

    cap = layout["capacitacion"]
    if clear_existing_merges:
        _unmerge_col_range(ws, 2, 3, cap["row_start"], cap["row_end"])
    ws.merge_cells(
        start_row=cap["row_start"],
        start_column=2,
        end_row=cap["row_end"],
        end_column=2,
    )
    ws.merge_cells(
        start_row=cap["row_start"],
        start_column=3,
        end_row=cap["row_end"],
        end_column=3,
    )
    ws.cell(cap["row_start"], 2).value = cap["label"]


def _write_atador_blocks(ws, layout: dict, clear_existing_merges: bool = True) -> None:
    """Write per-block data: col C key, day values, formulas, merges."""
    from openpyxl.utils import column_index_from_string as ci  # noqa: PLC0415

    for turn in layout["turns"].values():
        for block in turn["blocks"]:
            rs = block["row_start"]
            re_ = block["row_end"]
            items_by_day = block["items_by_day"]
            atador_key = block["atador_key"]

            # Col C: atador key
            if clear_existing_merges:
                _unmerge_col_range(ws, 3, 3, rs, re_)
            ws.merge_cells(start_row=rs, start_column=3, end_row=re_, end_column=3)
            ws.cell(rs, 3).value = atador_key

            # Write records row by row
            rows_data: list[dict[int, dict]] = [{} for _ in range(re_ - rs + 1)]
            for day_idx, day_items in enumerate(items_by_day):
                for row_offset, item in enumerate(day_items):
                    rows_data[row_offset][day_idx] = item

            for row_offset, day_map in enumerate(rows_data):
                row = rs + row_offset
                for day_idx, item in day_map.items():
                    d = DAY_DEFINITIONS[day_idx]
                    c_start = ci(d["start"])
                    c_end = ci(d["end"])
                    c_atado = ci(d["atado"])
                    c_telar = ci(d["telar"])
                    c_calif = ci(d["calif"])
                    c_5s = ci(d["five_s"])
                    c_merma = ci(d["merma"])

                    if item["hora_inicio"]:
                        ws.cell(row, c_start).value = to_excel_time(item["hora_inicio"])
                    if item["hora_fin"]:
                        ws.cell(row, c_end).value = to_excel_time(item["hora_fin"])
                    ws.cell(row, c_atado).value = item["atado"]
                    ws.cell(row, c_telar).value = item["telar"]
                    if item["calif"] is not None:
                        ws.cell(row, c_calif).value = item["calif"]
                    if item["five_s"] is not None:
                        ws.cell(row, c_5s).value = item["five_s"]
                    if item["merma"] is not None:
                        ws.cell(row, c_merma).value = item["merma"]

            # Duration + avg formulas
            _write_block_formulas(ws, block)

            # Merge CJ column over block
            cj = ci("CJ")
            if clear_existing_merges:
                _unmerge_col_range(ws, cj, cj, rs, re_)
            ws.merge_cells(start_row=rs, start_column=cj, end_row=re_, end_column=cj)

            # Merge avg columns
            for day in DAY_DEFINITIONS.values():
                for col_letter in [day["avg_time"], day["avg_calif"], day["avg_merma"]]:
                    c = ci(col_letter)
                    if clear_existing_merges:
                        _unmerge_col_range(ws, c, c, rs, re_)
                    ws.merge_cells(
                        start_row=rs, start_column=c, end_row=re_, end_column=c
                    )


def _write_block_formulas(ws, block: dict) -> None:
    """Duration per row + avg per block + CJ count."""
    from openpyxl.utils import column_index_from_string as ci  # noqa: PLC0415

    rs = block["row_start"]
    re_ = block["row_end"]

    for day in DAY_DEFINITIONS.values():
        sc = day["start"]
        ec = day["end"]
        dc = day["duration"]
        tc = day["telar"]
        avgt = day["avg_time"]
        ac = day["avg_calif"]
        fs = day["five_s"]
        avc = day["avg_calif"]
        mc = day["merma"]
        avm = day["avg_merma"]

        for row in range(rs, re_ + 1):
            ws.cell(
                row, ci(dc)
            ).value = f'=IF(OR({ec}{row}="",{sc}{row}=""),"",{ec}{row}-{sc}{row})'

        ws.cell(rs, ci(avgt)).value = (
            f'=IF(COUNT({tc}{rs}:{tc}{re_})=0,"",'
            f"SUM({dc}{rs}:{dc}{re_})/COUNT({tc}{rs}:{tc}{re_}))"
        )
        ws.cell(
            rs, ci(avc)
        ).value = f'=IF(COUNTA({ac}{rs}:{fs}{re_})=0,"",AVERAGE({ac}{rs}:{fs}{re_}))'
        ws.cell(
            rs, ci(avm)
        ).value = f'=IF(COUNT({mc}{rs}:{mc}{re_})=0,"",AVERAGE({mc}{rs}:{mc}{re_}))'

    # CJ: total count across all days
    count_parts = ",".join(
        f"{d['footer_count']}{rs}:{d['footer_count']}{re_}"
        for d in DAY_DEFINITIONS.values()
    )
    ws.cell(rs, 88).value = f"=COUNT({count_parts})"

    # CI min row (row_start + 1)
    if rs + 1 <= re_:
        min_parts = ",".join(d["avg_time"] + str(rs) for d in DAY_DEFINITIONS.values())
        ws.cell(rs + 1, 87).value = f"=MIN({min_parts})"  # CI=87


def _write_footer(ws, layout: dict) -> None:
    """Write footer row: B=SEMANA, C=week_num, CJ=ATADOS, count formulas."""
    from openpyxl.utils import column_index_from_string as ci  # noqa: PLC0415

    fr = layout["footer_row"]
    ds = layout["detail_start_row"]
    ld = layout["last_detail_row"]
    wn = layout.get("week_number", "")

    ws.cell(fr, 2).value = "SEMANA"
    ws.cell(fr, 3).value = wn
    ws.cell(fr, ci("CJ")).value = "ATADOS"

    footer_cols = []
    for day in DAY_DEFINITIONS.values():
        fc = day["footer_count"]
        formula = f"=COUNT({fc}{ds}:{fc}{ld})"
        ws.cell(fr, ci(fc)).value = formula
        footer_cols.append(f"{fc}{fr}")

    ws.cell(fr, 87).value = "=" + "+".join(footer_cols)  # CI=87


def _apply_ck_cu_styles(
    ws,
    section_top: int,
    data_rows: list[tuple[int, str]],
    last_detail_row: int | None = None,
) -> None:
    """Apply header labels, fills, borders, and number formats to CK-CU cols (89-99)."""
    st = _ck_styles()
    cols = range(89, 100)  # CK=89 .. CU=99

    # Row 3 relative (section_top + 2) = col-labels row → write CK-CU headers
    hdr_row = section_top + 2
    for i, col in enumerate(cols):
        c = ws.cell(hdr_row, col)
        c.value = _CK_HDR_LABELS[i]
        c.fill = st["hdr_fill"]
        c.font = st["hdr_font"]
        c.border = st["border"]
        c.alignment = st["center"]

    # Data rows — one per non-empty atador block
    for detalle_row, _ in data_rows:
        for col in cols:
            c = ws.cell(detalle_row, col)
            c.fill = st["data_fill"]
            c.font = st["body_font"]
            c.border = st["border"]
        # Time columns CN-CQ (92-95); CM is calidad and stays numeric.
        for col in (92, 93, 94, 95):
            ws.cell(detalle_row, col).number_format = "[h]:mm:ss"
        ws.cell(detalle_row, 96).number_format = "0.00"  # CR eficiencia %
        ws.cell(detalle_row, 97).number_format = "0.00"  # CS calidad %
        ws.cell(detalle_row, 98).number_format = "0.00"  # CT merma

    # Aggregate rows live at fixed offsets from the first summary row (row 4 within the section),
    # even when the first populated atador appears later.
    if data_rows:
        for agg_row in _ck_reserved_rows(section_top, last_detail_row):
            for col in cols:
                c = ws.cell(agg_row, col)
                try:
                    rgb = c.fill.fgColor.rgb if c.fill else None
                except Exception:
                    rgb = None
                if rgb in (None, "00000000", "FFFFFFFF"):
                    c.fill = st["agg_fill"]
                c.font = st["body_font"]
                c.border = st["border"]


def _write_ck_cu_formulas(
    ws, layout: dict, section_top: int, name_map: dict
) -> list[tuple[int, str]]:
    """
    Mirror of PHP writeCkCuFormulas.
    Summary rows start at section_top + 3 (1-indexed = row 4 relative to section start).
    Returns list of (detalle_row, atador_key) for each non-empty block,
    used by update_semana_sheet to write the correct DETALLE references.
    """
    last_detail_row = layout.get("last_detail_row")
    reserved_rows = set(_ck_reserved_rows(section_top, last_detail_row))
    summary_row = section_top + 3
    actual_summary_rows: list[tuple[int, str]] = []

    for turn in layout["turns"].values():
        for block in turn["blocks"]:
            while summary_row in reserved_rows:
                summary_row += 1
            row = summary_row
            summary_row += 1
            key = str(block["atador_key"] or "").strip()
            bs = block["row_start"]

            if not key:
                # Clear CK..CU for this row
                for col in range(89, 100):  # CK=89..CU=99
                    ws.cell(row, col).value = None
                continue

            actual_summary_rows.append((row, key))  # (detalle_row, atador_key)
            name = name_map.get(key, key)

            ws.cell(row, 89).value = f"=C{bs}"  # CK
            ws.cell(row, 90).value = name  # CL
            ws.cell(row, 91).value = (
                "=IFERROR(AVERAGE("
                + ",".join(f"{c}{bs}" for c in AVG_CALIF_COLS)
                + '),"")'
            )  # CM
            ws.cell(row, 92).value = (
                "=IFERROR(AVERAGE("
                + ",".join(f"{c}{bs}" for c in AVG_TIME_COLS)
                + '),"")'
            )  # CN
            ws.cell(row, 97).value = f'=IFERROR(CM{row}*100/10,"")'  # CS
            ws.cell(row, 98).value = (
                "=IFERROR(AVERAGE("
                + ",".join(f"{c}{bs}" for c in AVG_MERMA_COLS)
                + '),"")'
            )  # CT
            ws.cell(row, 99).value = f"=CL{row}"  # CU

    if not actual_summary_rows:
        _apply_ck_cu_styles(ws, section_top, actual_summary_rows, last_detail_row)
        return actual_summary_rows

    cn_refs = ",".join(f"CN{r}" for r, _ in actual_summary_rows)
    for row, _ in actual_summary_rows:
        ws.cell(row, 93).value = f'=IFERROR(MIN({cn_refs}),"")'  # CO
        ws.cell(row, 94).value = f'=IFERROR(CN{row}-CO{row},"")'  # CP
        ws.cell(row, 95).value = f'=IFERROR(CO{row}-CP{row},"")'  # CQ
        ws.cell(row, 96).value = f'=IFERROR(CQ{row}*100/CO{row},"")'  # CR

    _apply_ck_cu_styles(ws, section_top, actual_summary_rows, last_detail_row)
    return actual_summary_rows


def _ck_reserved_rows(
    section_top: int, last_detail_row: int | None = None
) -> tuple[int, ...]:
    first_summary_row = section_top + 3
    rows = (
        first_summary_row + 7,
        first_summary_row + 9,
        first_summary_row + 10,
    )
    if last_detail_row is None:
        return rows
    return tuple(row for row in rows if row <= last_detail_row)


# ---------------------------------------------------------------------------
# SEMANA XX sheet helpers
# ---------------------------------------------------------------------------


def update_semana_sheet(
    wb,
    week_num: int,
    actual_summary_rows: list[tuple[int, str]],
    name_map: dict[str, str],
) -> None:
    """
    Create or update the SEMANA XX sheet.

    - Existing sheet → update B2, B4:C8 (atador key/name) and D4:K8 + L4:L8
      (DETALLE row references that shift when DETALLE grows). In-sheet formulas
      in rows 11,13,16-22 use relative refs and never need changing. Manual data
      in rows 26-51 is never touched.
    - New sheet → clone the last existing SEMANA sheet (for styles/structure),
      clear rows 26-51, then update B2 + B4:K8 + L4:L8.
    - Brand-new file with no existing SEMANA → create sheet from scratch via
      _write_semana_structure.

    actual_summary_rows: list of (detalle_row, atador_key) for each non-empty
                         atador block, in layout order (max 5 used for rows 4-8).
    """
    from openpyxl.utils import get_column_letter  # noqa: PLC0415

    sheet_name = f"SEMANA {week_num:02d}"
    semana_re = re.compile(r"^SEMANA\s+(\d+)$", re.IGNORECASE)

    # Collect existing SEMANA sheets sorted by week number
    existing_semana: dict[int, str] = {}
    for sname in wb.sheetnames:
        m = semana_re.match(sname.strip())
        if m:
            existing_semana[int(m.group(1))] = sname

    is_new = sheet_name not in wb.sheetnames
    clone_succeeded = False
    if is_new:
        if existing_semana:
            src_name = existing_semana[max(existing_semana.keys())]
            try:
                ws_sem = wb.copy_worksheet(wb[src_name])
                ws_sem.title = sheet_name
                clone_succeeded = True
            except Exception:  # noqa: BLE001
                # copy_worksheet can fail on sheets with charts; create fresh
                ws_sem = wb.create_sheet(sheet_name)
        else:
            ws_sem = wb.create_sheet(sheet_name)
        # Clear manual section (rows 26-51) — new sheet starts with clean slate
        for row in range(26, 52):
            for col in range(1, 20):
                ws_sem.cell(row, col).value = None
    else:
        ws_sem = wb[sheet_name]

    # B2: week label (always update so cloned/stale sheets get correct week)
    ws_sem["B2"] = f"SEMANA {week_num}"

    # Take up to 5 actual summary rows → SEMANA rows 4-8
    atador_items = actual_summary_rows[:5]

    # Columns D-K map to DETALLE CK-CU: CM=D, CN=E, CO=F, CP=G, CQ=H, CR=I, CS=J, CT=K
    detalle_cols = ["CM", "CN", "CO", "CP", "CQ", "CR", "CS", "CT"]

    for i, (detalle_row, atador_key) in enumerate(atador_items):
        sheet_row = 4 + i

        # B: atador key (numeric if possible, otherwise string)
        try:
            ws_sem.cell(sheet_row, 2).value = int(atador_key)
        except (ValueError, TypeError):
            ws_sem.cell(sheet_row, 2).value = atador_key

        # C: atador name from name_map (fall back to key if not found)
        ws_sem.cell(sheet_row, 3).value = name_map.get(atador_key, atador_key)

        # D-K: DETALLE CK-CU references (update whenever DETALLE rows shift)
        for col_offset, dc in enumerate(detalle_cols):
            ws_sem.cell(
                sheet_row, 4 + col_offset
            ).value = f'=IFERROR(DETALLE!{dc}{detalle_row},"")'

        # L: echo of C (nombre eco)
        ws_sem.cell(sheet_row, 12).value = f"=C{sheet_row}"

        # M: 5S manual score — preserve existing value; default 100 only if blank
        if ws_sem.cell(sheet_row, 13).value is None:
            ws_sem.cell(sheet_row, 13).value = 100

    # Clear rows beyond actual atadors (in case previous export had more)
    for i in range(len(atador_items), 5):
        sheet_row = 4 + i
        for col in range(2, 13):  # B..L
            ws_sem.cell(sheet_row, col).value = None

    # For new sheets without a successful clone: write the static OEE skeleton
    if is_new and not clone_succeeded:
        _write_semana_structure(ws_sem)


def _write_semana_structure(ws_sem) -> None:
    """
    Write the static skeleton of a SEMANA sheet (labels + in-sheet formulas).
    Called only when creating a brand-new sheet that couldn't be cloned.
    """
    from openpyxl.utils import get_column_letter  # noqa: PLC0415

    # Row 11: avg merma across atadors
    ws_sem["K11"] = '=IFERROR(AVERAGE(K4:K8),"")'

    # Row 13: promedio general OEE (averages OEE row 22 across cols D-H)
    ws_sem["C13"] = "PROMEDIO GRAL."
    ws_sem["D13"] = '=IFERROR(AVERAGE(D22:H22),"")'

    # Row 16: atador names
    ws_sem["C16"] = "ATADOR"
    for col, src_row in zip(range(4, 9), range(4, 9)):
        ws_sem.cell(16, col).value = f"=C{src_row}"

    # Row 17: eficiencia atador (CR → col I in SEMANA rows 4-8)
    ws_sem["C17"] = "EFIC. ATADOR"
    for col, src_row in zip(range(4, 9), range(4, 9)):
        ws_sem.cell(17, col).value = f'=IFERROR(I{src_row},"")'

    # Row 18: efic. x auxiliar — label only (no formulas in existing file)
    ws_sem["C18"] = "EFIC. X AUXILIAR"

    # Row 19: calidad/5S seguridad
    ws_sem["C19"] = "CALIDAD/5S SEGURIDAD"
    for col, src_row in zip(range(4, 9), range(4, 9)):
        ws_sem.cell(19, col).value = f'=IFERROR(AVERAGE(J{src_row},M{src_row}),"")'

    # Row 20: merma promedio
    ws_sem["B20"] = '=IFERROR(MIN(K4:K8),"")'
    ws_sem["C20"] = "MERMA (PROMEDIO)"
    for col, src_row in zip(range(4, 9), range(4, 9)):
        ws_sem.cell(20, col).value = f'=IFERROR(K{src_row},"")'

    # Row 21: % merma
    ws_sem["C21"] = "% X MERMA"
    for col in range(4, 9):
        cl = get_column_letter(col)
        ws_sem.cell(21, col).value = f'=IFERROR($B$20*100/{cl}20,"")'

    # Row 22: OEE
    ws_sem["C22"] = "OEE"
    for col in range(4, 9):
        cl = get_column_letter(col)
        ws_sem.cell(22, col).value = f'=IFERROR({cl}21*{cl}19*{cl}17/1000000,"")'


# ---------------------------------------------------------------------------
# Main export orchestrator
# ---------------------------------------------------------------------------


def get_weeks_in_range(week_start: date, week_end: date) -> list[date]:
    """Return list of Monday dates from week_start to week_end inclusive."""
    weeks = []
    d = week_start
    while d <= week_end:
        weeks.append(d)
        d += timedelta(days=7)
    return weeks


def _update_parsed_after_insert(
    parsed: dict, insert_top: int, section_rows: int, week_num: int, new_section: dict
) -> dict:
    """
    Update the in-memory `parsed` dict after inserting rows at `insert_top`.
    Shifts all existing sections at or below `insert_top` by `+section_rows` and
    registers the new section. Avoids a full re-scan of DETALLE.
    """
    new_sections: list[dict] = []
    new_map: dict[int, dict] = {}
    for sec in parsed["sections"]:
        if sec["top"] >= insert_top:
            shifted = {
                **sec,
                "top": sec["top"] + section_rows,
                "footer": sec["footer"] + section_rows,
            }
            new_sections.append(shifted)
            new_map[shifted["week"]] = shifted
        else:
            new_sections.append(sec)
            new_map[sec["week"]] = sec
    new_sections.append(new_section)
    new_map[week_num] = new_section
    new_sections.sort(key=lambda s: s["top"])
    return {"sections": new_sections, "map": new_map}


def run(args: argparse.Namespace) -> None:
    from openpyxl import load_workbook  # noqa: PLC0415

    _set_run_phase("inicio")

    t0 = time.time()
    status_file = args.status_file
    file_path = args.file_path

    write_status(status_file, "procesando", "Iniciando...")

    week_start_date = date.fromisoformat(args.week_start)
    week_end_date = date.fromisoformat(args.week_end)
    weeks = get_weeks_in_range(week_start_date, week_end_date)
    iso_year = assert_single_iso_year(weeks)

    # 1. Copy to local temp
    _set_run_phase("copiar_entrada")
    tmp_dir = tempfile.gettempdir()
    local_temp = os.path.join(tmp_dir, f"oee_py_{args.token}.xlsx")
    copy_oee_source_to_local_temp(file_path, local_temp, status_file, t0)
    write_status(
        status_file, "procesando", "Archivo copiado localmente", time.time() - t0
    )

    # 2. Query DB
    _set_run_phase("conectar_bd")
    conn = get_db_connection(args)
    _set_run_phase("consultar_bd")
    raw_rows = query_records(conn, args.week_start, args.week_end)
    conn.close()

    grouped = group_records_by_week(raw_rows)

    write_status(
        status_file,
        "procesando",
        "Datos cargados, actualizando Excel...",
        time.time() - t0,
    )

    # 3. Load workbook
    _set_run_phase("cargar_excel")
    wb = load_workbook(local_temp, keep_vba=False, data_only=False)
    ws = wb["DETALLE"]

    wby = resolve_workbook_year(wb)
    if wby is not None and wby != iso_year:
        raise RuntimeError(
            f"El archivo OEE corresponde al año {wby}; no se pueden mezclar semanas del año ISO {iso_year}."
        )

    parsed0 = parse_detalle_sections(ws)
    normalize_detalle_footer_weeks(ws, parsed0["sections"])

    # Capture a style template from the first existing section (used for brand-new inserts)
    _template_snap: dict | None = None
    if parsed0["sections"]:
        _template_snap = snapshot_section_styles(ws, parsed0["sections"][0])

    # Índice de secciones: se reutiliza entre iteraciones; solo se vuelve a parsear tras insert/resize.
    parsed: dict = parsed0

    # 4. Process each week
    _set_run_phase("procesar_semanas")
    for week_monday in weeks:
        week_key = str(week_monday)
        week_num = week_monday.isocalendar()[1]

        week_raw = grouped.get(week_key, [])
        week_records = prepare_records(week_raw, week_monday)

        # Build name_map
        name_map: dict[str, str] = {}
        for rec in week_records:
            k = rec["atador_key"]
            if k and k not in name_map:
                for r in week_raw:
                    if resolve_atador_key(r) == k:
                        nm = str(r.get("NomTejedor") or "").strip()
                        if nm:
                            name_map[k] = nm
                        break

        existing = parsed["map"].get(week_num)

        if existing:
            layout = build_layout(week_records, section_top_row=existing["top"])
            layout["week_number"] = week_num
            desired_total = layout["footer_row"] - existing["top"] + 1

            # PERF: skip expensive style snapshot/restore when section size unchanged
            if desired_total != existing["rows"]:
                snap = snapshot_section_styles(ws, existing)
                existing = resize_section(ws, existing, desired_total)
                layout = build_layout(week_records, section_top_row=existing["top"])
                layout["week_number"] = week_num
                clear_section_values(ws, existing, already_unmerged=True)
                apply_section_styles(ws, existing, snap, layout)
                parsed = parse_detalle_sections(ws)
            else:
                clear_section_values(ws, existing)

            actual_summary_rows = write_section_data(
                ws,
                existing,
                layout,
                week_monday,
                name_map,
                clear_existing_merges=False,
            )
        else:
            insert_top = resolve_insert_top(parsed["sections"], week_num)

            layout_tmp = build_layout(week_records, section_top_row=insert_top)
            layout_tmp["week_number"] = week_num
            section_rows = layout_tmp["footer_row"] - insert_top + 1

            ws.insert_rows(insert_top, section_rows)
            sync_merged_ranges_after_row_insert(ws, insert_top, section_rows)
            _translate_shifted_formulas(ws, insert_top, section_rows)

            new_section = {
                "top": insert_top,
                "footer": insert_top + section_rows - 1,
                "rows": section_rows,
                "week": week_num,
            }
            layout = build_layout(week_records, section_top_row=insert_top)
            layout["week_number"] = week_num

            if _template_snap:
                apply_section_styles(ws, new_section, _template_snap, layout)

            actual_summary_rows = write_section_data(
                ws,
                new_section,
                layout,
                week_monday,
                name_map,
                clear_existing_merges=False,
            )
            parsed = _update_parsed_after_insert(
                parsed, insert_top, section_rows, week_num, new_section
            )

        # Create/update SEMANA XX sheet with OEE formulas referencing DETALLE rows
        update_semana_sheet(wb, week_num, actual_summary_rows, name_map)

        write_status(
            status_file,
            "procesando",
            f"Semana {week_num} actualizada ({round(time.time() - t0)}s)",
            time.time() - t0,
        )

    normalize_detalle_footer_weeks(ws, parsed["sections"])
    normalize_detalle_visual_weeks(ws, parsed["sections"])
    patch_semana_sheets_week_labels(wb, iso_year)

    # 5. Save to temp output
    _set_run_phase("guardar_excel")
    out_temp = os.path.join(tmp_dir, f"oee_py_out_{args.token}.xlsx")
    write_status(status_file, "procesando", "Guardando archivo...", time.time() - t0)
    wb.save(out_temp)
    wb.close()

    # 6. Copy back to network (retry on file-lock errors common on shares)
    _set_run_phase("copiar_salida")
    write_status(
        status_file,
        "procesando",
        "Copiando resultado al archivo de red…",
        time.time() - t0,
    )
    alternate_saved = copy_oee_result_to_destination(
        out_temp, file_path, status_file, t0, args.token
    )

    # Cleanup temps
    for f in [local_temp, out_temp]:
        try:
            os.remove(f)
        except OSError:
            pass

    elapsed = round(time.time() - t0, 1)
    if alternate_saved:
        write_status(
            status_file,
            "completado",
            f"Completado en {elapsed}s. El archivo principal estaba bloqueado; "
            f"el Excel actualizado quedó en: {alternate_saved}. "
            "Cierra OEE_ATADORES.xlsx en Excel (en todas las PCs) y sustituye el archivo "
            "renombrando esta copia al nombre original, o vuelve a ejecutar el export.",
            elapsed,
        )
    else:
        write_status(status_file, "completado", f"Completado en {elapsed}s", elapsed)


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------


def main() -> None:
    args = parse_args()

    # Optional log file
    if getattr(args, "log_file", None):
        import logging  # noqa: PLC0415

        logging.basicConfig(
            filename=args.log_file,
            level=logging.INFO,
            format="%(asctime)s %(levelname)s %(message)s",
        )

    t0 = time.time()
    try:
        run(args)
    except BaseException as exc:  # noqa: BLE001
        elapsed = round(time.time() - t0, 1)
        tb = traceback.format_exc()
        print(tb, file=sys.stderr, end="")
        msg = f"fase={_RUN_PHASE} | {_format_exception_detail(exc)}"
        if len(msg) < 100 and tb:
            msg = f"{msg}\n{tb}"[:4000]
        msg = msg[:4000]
        write_status(args.status_file, "error", msg, elapsed)
        print(f"ERROR: {msg}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
